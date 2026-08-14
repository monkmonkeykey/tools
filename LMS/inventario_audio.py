#!/usr/bin/env python3
"""
Genera un inventario Excel con las características técnicas de archivos de audio.

CONFIGURACIÓN:
    Modifica CARPETA_PRINCIPAL y ARCHIVO_EXCEL en la sección
    "RUTAS Y OPCIONES" de este mismo archivo.

Ejecución:
    python3 inventario_audio_rutas_codigo.py

Dependencias:
    python3 -m pip install mutagen openpyxl
"""

from __future__ import annotations

import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any


# =============================================================================
# RUTAS Y OPCIONES: MODIFICA ÚNICAMENTE ESTA SECCIÓN
# =============================================================================

# Carpeta principal que contiene una subcarpeta por cada obra.
# En macOS, un ejemplo puede ser:
# CARPETA_PRINCIPAL = "/Users/tu_usuario/Desktop/Obras"
CARPETA_PRINCIPAL = Path(
    "/Users/josue/Library/CloudStorage/"
    "ProtonDrive-studio@luzmariasanchez.com-folder/"
    "Archivos restaurados 21 03 2026, 09:05:46/"
    "BECAS Y CONVOCATORIAS/"
    "2026 Trondheim_kommune_2026_Sound_Works_1997-2024_Saksnr_29337/"
    "02_Original_Sound_Works"
)

# Ruta completa y nombre del Excel que se generará.
# La carpeta de destino se crea automáticamente si no existe.
ARCHIVO_EXCEL = "/Users/josue/Downloads/"

# Cómo determinar el nombre de la obra:
# "first"  = primera subcarpeta dentro de CARPETA_PRINCIPAL. Recomendado.
# "parent" = carpeta inmediata que contiene cada archivo de audio.
MODO_OBRA = "first"

# =============================================================================

try:
    from mutagen import File as MutagenFile
    from mutagen import MutagenError
except ImportError as exc:
    raise SystemExit(
        "Falta la dependencia 'mutagen'. Instálala con:\n"
        "python3 -m pip install mutagen openpyxl"
    ) from exc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit(
        "Falta la dependencia 'openpyxl'. Instálala con:\n"
        "python3 -m pip install mutagen openpyxl"
    ) from exc


EXTENSIONES_AUDIO = {
    ".wav",
    ".wave",
    ".aif",
    ".aiff",
    ".aifc",
    ".mp3",
    ".flac",
    ".m4a",
    ".mp4",
    ".aac",
    ".ogg",
    ".oga",
    ".opus",
    ".wma",
}

ENCABEZADOS = [
    "Obra",
    "Subcarpeta relativa",
    "Nombre del archivo",
    "Extensión",
    "Formato detectado",
    "MIME",
    "Códec / descripción",
    "Duración",
    "Duración (segundos)",
    "Sample rate (Hz)",
    "Sample rate (kHz)",
    "Canales",
    "Configuración de canales",
    "Profundidad de bits",
    "Bitrate (kbps)",
    "Tamaño (MB)",
    "Fecha de modificación",
    "Ruta relativa",
    "Estado",
    "Observaciones / error",
]


# -----------------------------------------------------------------------------
# Utilidades
# -----------------------------------------------------------------------------

def valor_atributo(objeto: Any, nombre: str, valor_predeterminado: Any = None) -> Any:
    """Obtiene un atributo sin fallar si el formato no lo proporciona."""
    return getattr(objeto, nombre, valor_predeterminado)


def redondear(valor: Any, decimales: int = 3) -> float | None:
    if valor is None:
        return None
    try:
        return round(float(valor), decimales)
    except (TypeError, ValueError):
        return None


def descripcion_canales(canales: int | None) -> str | None:
    if canales is None:
        return None
    equivalencias = {
        1: "Mono",
        2: "Estéreo",
        4: "4 canales",
        6: "5.1 o 6 canales",
        8: "7.1 u 8 canales",
    }
    return equivalencias.get(canales, f"{canales} canales")


def detectar_obra(carpeta_raiz: Path, archivo: Path, modo: str) -> str:
    """
    first: primera subcarpeta debajo de la raíz.
    parent: carpeta inmediata que contiene el archivo.
    """
    if modo == "parent":
        return archivo.parent.name if archivo.parent != carpeta_raiz else "(raíz)"

    padre_relativo = archivo.parent.relative_to(carpeta_raiz)
    return padre_relativo.parts[0] if padre_relativo.parts else "(raíz)"


def describir_codec(audio: Any, info: Any, extension: str) -> str:
    """Construye una descripción legible con la información disponible."""
    descripcion = valor_atributo(info, "codec_description")
    if descripcion:
        return str(descripcion)

    codec = valor_atributo(info, "codec")
    if codec:
        return str(codec)

    if extension == ".mp3":
        partes = []
        version = valor_atributo(info, "version")
        layer = valor_atributo(info, "layer")
        modo = valor_atributo(info, "mode")
        bitrate_mode = valor_atributo(info, "bitrate_mode")

        if version is not None:
            partes.append(f"MPEG {version}")
        if layer is not None:
            partes.append(f"Layer {layer}")
        if modo is not None:
            partes.append(f"modo {modo}")
        if bitrate_mode is not None:
            nombre_modo = getattr(bitrate_mode, "name", str(bitrate_mode))
            partes.append(str(nombre_modo))

        if partes:
            return ", ".join(partes)

    nombre_info = type(info).__name__
    if nombre_info and nombre_info != "NoneType":
        return nombre_info.replace("StreamInfo", "").replace("Info", "")

    return type(audio).__name__


def analizar_archivo(carpeta_raiz: Path, archivo: Path, modo_obra: str) -> dict[str, Any]:
    extension = archivo.suffix.lower()
    ruta_relativa = archivo.relative_to(carpeta_raiz)
    subcarpeta_relativa = archivo.parent.relative_to(carpeta_raiz)

    datos: dict[str, Any] = {
        "obra": detectar_obra(carpeta_raiz, archivo, modo_obra),
        "subcarpeta": str(subcarpeta_relativa) if subcarpeta_relativa.parts else ".",
        "nombre": archivo.name,
        "extension": extension,
        "formato": None,
        "mime": None,
        "codec": None,
        "duracion_excel": None,
        "duracion_segundos": None,
        "sample_rate_hz": None,
        "sample_rate_khz": None,
        "canales": None,
        "configuracion_canales": None,
        "bits": None,
        "bitrate_kbps": None,
        "tamano_mb": None,
        "fecha_modificacion": None,
        "ruta_relativa": str(ruta_relativa),
        "estado": "OK",
        "observaciones": "",
    }

    try:
        estadisticas = archivo.stat()
        datos["tamano_mb"] = round(estadisticas.st_size / (1024 * 1024), 3)
        datos["fecha_modificacion"] = datetime.fromtimestamp(estadisticas.st_mtime)

        audio = MutagenFile(archivo)
        if audio is None or not hasattr(audio, "info"):
            raise ValueError("Mutagen no reconoció el contenido del archivo")

        info = audio.info
        duracion = redondear(valor_atributo(info, "length"), 3)
        sample_rate = valor_atributo(info, "sample_rate")
        canales = valor_atributo(info, "channels")
        bits = valor_atributo(info, "bits_per_sample")
        bitrate = valor_atributo(info, "bitrate")

        datos["formato"] = type(audio).__name__
        mime = valor_atributo(audio, "mime")
        datos["mime"] = ", ".join(mime) if isinstance(mime, (list, tuple)) else mime
        datos["codec"] = describir_codec(audio, info, extension)
        datos["duracion_segundos"] = duracion
        # Excel representa las duraciones como fracciones de un día.
        datos["duracion_excel"] = duracion / 86400 if duracion is not None else None
        datos["sample_rate_hz"] = int(sample_rate) if sample_rate is not None else None
        datos["sample_rate_khz"] = round(sample_rate / 1000, 3) if sample_rate else None
        datos["canales"] = int(canales) if canales is not None else None
        datos["configuracion_canales"] = descripcion_canales(datos["canales"])
        datos["bits"] = int(bits) if bits is not None else None
        datos["bitrate_kbps"] = round(bitrate / 1000, 2) if bitrate else None

        if extension == ".mp3" and datos["bits"] is None:
            datos["observaciones"] = (
                "La profundidad de bits no aplica directamente a MP3 como en audio PCM."
            )

    except (MutagenError, OSError, ValueError, TypeError) as exc:
        datos["estado"] = "ERROR"
        datos["observaciones"] = str(exc)

    return datos


def buscar_archivos(carpeta_raiz: Path) -> list[Path]:
    archivos = [
        ruta
        for ruta in carpeta_raiz.rglob("*")
        if ruta.is_file() and ruta.suffix.lower() in EXTENSIONES_AUDIO
    ]
    return sorted(archivos, key=lambda p: str(p.relative_to(carpeta_raiz)).lower())


# -----------------------------------------------------------------------------
# Excel
# -----------------------------------------------------------------------------

def aplicar_estilo_encabezado(hoja, rango_fila: int, total_columnas: int) -> None:
    relleno = PatternFill("solid", fgColor="1F4E78")
    fuente = Font(color="FFFFFF", bold=True)

    for celda in hoja[rango_fila][:total_columnas]:
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ajustar_anchos(hoja, maximo: int = 42) -> None:
    for columna in hoja.columns:
        letra = columna[0].column_letter
        longitud = 0
        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            longitud = max(longitud, len(valor))
        hoja.column_dimensions[letra].width = min(max(longitud + 2, 10), maximo)


def crear_excel(datos: list[dict[str, Any]], salida: Path, carpeta_raiz: Path) -> None:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Archivos"

    hoja.append(ENCABEZADOS)

    for item in datos:
        hoja.append(
            [
                item["obra"],
                item["subcarpeta"],
                item["nombre"],
                item["extension"],
                item["formato"],
                item["mime"],
                item["codec"],
                item["duracion_excel"],
                item["duracion_segundos"],
                item["sample_rate_hz"],
                item["sample_rate_khz"],
                item["canales"],
                item["configuracion_canales"],
                item["bits"],
                item["bitrate_kbps"],
                item["tamano_mb"],
                item["fecha_modificacion"],
                item["ruta_relativa"],
                item["estado"],
                item["observaciones"],
            ]
        )

    aplicar_estilo_encabezado(hoja, 1, len(ENCABEZADOS))
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    hoja.sheet_view.showGridLines = False

    # Formatos numéricos.
    for celda in hoja["H"][1:]:
        celda.number_format = "[h]:mm:ss.000"
    for celda in hoja["I"][1:]:
        celda.number_format = "0.000"
    for celda in hoja["J"][1:]:
        celda.number_format = "0"
    for celda in hoja["K"][1:]:
        celda.number_format = "0.000"
    for celda in hoja["O"][1:]:
        celda.number_format = "0.00"
    for celda in hoja["P"][1:]:
        celda.number_format = "0.000"
    for celda in hoja["Q"][1:]:
        celda.number_format = "yyyy-mm-dd hh:mm:ss"

    # Tabla estructurada, solo cuando existen registros.
    if datos:
        tabla = Table(displayName="InventarioAudio", ref=hoja.dimensions)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja.add_table(tabla)

    ajustar_anchos(hoja)
    hoja.column_dimensions["R"].width = 55
    hoja.column_dimensions["T"].width = 55

    # Resumen por obra.
    resumen = defaultdict(lambda: {"archivos": 0, "duracion": 0.0, "tamano": 0.0, "errores": 0})
    for item in datos:
        registro = resumen[item["obra"]]
        registro["archivos"] += 1
        registro["duracion"] += item["duracion_segundos"] or 0
        registro["tamano"] += item["tamano_mb"] or 0
        if item["estado"] != "OK":
            registro["errores"] += 1

    hoja_resumen = libro.create_sheet("Resumen por obra")
    hoja_resumen.append(
        ["Obra", "Número de archivos", "Duración total", "Duración total (segundos)", "Tamaño total (MB)", "Errores"]
    )

    for obra in sorted(resumen, key=str.lower):
        registro = resumen[obra]
        hoja_resumen.append(
            [
                obra,
                registro["archivos"],
                registro["duracion"] / 86400,
                round(registro["duracion"], 3),
                round(registro["tamano"], 3),
                registro["errores"],
            ]
        )

    aplicar_estilo_encabezado(hoja_resumen, 1, 6)
    hoja_resumen.freeze_panes = "A2"
    hoja_resumen.sheet_view.showGridLines = False
    for celda in hoja_resumen["C"][1:]:
        celda.number_format = "[h]:mm:ss.000"
    for celda in hoja_resumen["D"][1:]:
        celda.number_format = "0.000"
    for celda in hoja_resumen["E"][1:]:
        celda.number_format = "0.000"

    if resumen:
        tabla_resumen = Table(displayName="ResumenObras", ref=hoja_resumen.dimensions)
        tabla_resumen.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja_resumen.add_table(tabla_resumen)

    ajustar_anchos(hoja_resumen)

    # Información de ejecución.
    hoja_info = libro.create_sheet("Información")
    hoja_info.append(["Dato", "Valor"])
    hoja_info.append(["Carpeta revisada", str(carpeta_raiz)])
    hoja_info.append(["Fecha de ejecución", datetime.now()])
    hoja_info.append(["Archivos encontrados", len(datos)])
    hoja_info.append(["Extensiones consideradas", ", ".join(sorted(EXTENSIONES_AUDIO))])
    aplicar_estilo_encabezado(hoja_info, 1, 2)
    hoja_info["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
    hoja_info.sheet_view.showGridLines = False
    ajustar_anchos(hoja_info, maximo=80)

    salida.parent.mkdir(parents=True, exist_ok=True)
    libro.save(salida)


# -----------------------------------------------------------------------------
# Programa principal
# -----------------------------------------------------------------------------

def main() -> int:
    carpeta_texto = str(CARPETA_PRINCIPAL).strip().strip('"').strip("'")
    salida_texto = str(ARCHIVO_EXCEL).strip().strip('"').strip("'")

    carpeta_raiz = Path(carpeta_texto).expanduser().resolve()
    salida = Path(salida_texto).expanduser().resolve()

    if MODO_OBRA not in {"first", "parent"}:
        print(
            'ERROR: MODO_OBRA debe ser "first" o "parent".',
            file=sys.stderr,
        )
        return 1

    if not carpeta_raiz.exists():
        print(f"ERROR: La carpeta no existe: {carpeta_raiz}", file=sys.stderr)
        return 1
    if not carpeta_raiz.is_dir():
        print(f"ERROR: La ruta no es una carpeta: {carpeta_raiz}", file=sys.stderr)
        return 1

    if salida.suffix.lower() != ".xlsx":
        salida = salida.with_suffix(".xlsx")

    archivos = buscar_archivos(carpeta_raiz)
    print(f"Carpeta revisada: {carpeta_raiz}")
    print(f"Archivos de audio encontrados: {len(archivos)}")

    datos = []
    for indice, archivo in enumerate(archivos, start=1):
        print(f"[{indice}/{len(archivos)}] {archivo.relative_to(carpeta_raiz)}")
        datos.append(analizar_archivo(carpeta_raiz, archivo, MODO_OBRA))

    crear_excel(datos, salida, carpeta_raiz)

    errores = sum(1 for item in datos if item["estado"] != "OK")
    print("\nProceso terminado.")
    print(f"Excel generado: {salida}")
    print(f"Registros con error: {errores}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())